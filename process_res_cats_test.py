# -*- coding: utf-8 -*-
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import datetime as dt
import re
import glob
from difflib import get_close_matches

def op_res_cats_files(path,company,op_res_cats_report):
    templates=[]
    op_template=pd.read_excel('Op_Res_Cats/Datamodel OpCats-ResCats-LoV.xlsx',pd.ExcelFile('Op_Res_Cats/Datamodel OpCats-ResCats-LoV.xlsx').sheet_names[0])
    res_template=pd.read_excel('Op_Res_Cats/Datamodel OpCats-ResCats-LoV.xlsx',pd.ExcelFile('Op_Res_Cats/Datamodel OpCats-ResCats-LoV.xlsx').sheet_names[1])
    templates.append(op_template)
    templates.append(res_template)
    sheets=[]
    res=pd.DataFrame()
    ops=pd.DataFrame()
    sheetnames=['RES.CAT.','RES CAT','ResCat','OP.CAT.','OP CAT','OpCat']
    unmatched_fields_res=[]
    unmatched_fields_ops=[]
    for j in range(len(glob.glob(path+'/*'))):
        if glob.glob(path+'/*')[j].endswith(('.xls','.xlsx')):
            for k in range(len(list(set(pd.ExcelFile(glob.glob(path+'*')[j]).sheet_names).intersection(sheetnames)))):
                sheets.append(pd.read_excel(glob.glob(path+'*')[j],list(set(pd.ExcelFile(glob.glob(path+'*')[j]).sheet_names).intersection(sheetnames))[k]))
            #files=pd.read_excel(glob.glob(path+'*')[j],sheet_name=None)
            #for frame in files.keys():
            #   sheets.append(files[frame])
        elif glob.glob(path+'/*')[j].endswith('.csv'):
            sheets.append(pd.read_csv(glob.glob(path+'*')[j],sep=";",encoding='ISO-8859-1'))
        else:
            None
    ###check fsheetnames
    if len(sheets)==0:
        
        print('','Check the sheetnames in the input files','The values should be: ResCat or/and OpCat',sep='\n',file=open(op_res_cats_report +'Mismatched_fields.txt','a',encoding='utf8'))
    else: 
        for j in range(len(sheets)):
            sheets[j].fillna('',inplace=True)
            for k in range(len(sheets[j].columns)):
                sheets[j].iloc[:,k]=pd.Series(np.where(sheets[j].iloc[:,k].str.contains('None'),'- None -',sheets[j].iloc[:,k]))
                sheets[j].iloc[:,k]=sheets[j].iloc[:,k].apply(lambda x: x.strip() if type(x)==str else x)
            #res
            if (sheets[j].columns.str.contains('ResCat',case=False).any()):
                res=sheets[j]
            else:
                None
            #ops
            if sheets[j].columns.str.contains('OpCat',case=False).any():
                ops=sheets[j]
            else:
                None
        
        if len(res)>0:
            wrong_Tier1=res.loc[~res.iloc[:,2].isin(templates[1].iloc[:,2]) & res.iloc[:,3].isin(templates[1].iloc[:,3])].drop_duplicates()
            wrong_Tier1['Wrong Tier'] ='Tier 1'
            wrong_Tier1['Possible match Tier 1']=wrong_Tier1.iloc[:,2].apply(lambda x: x if pd.isnull(x) or type(x)==float or type(x)==int else get_close_matches(x,templates[1].iloc[:,2].astype(str).unique().tolist(),1)).apply(lambda x: 'Not Found' if len(x)==0 or pd.isnull(x) or type(x)==float or type(x)==int else ''.join(x))
            wrong_Tier1['Possible match Tier 2']=np.nan
            wrong_Tier2=res.loc[~res.iloc[:,3].isin(templates[1].iloc[:,3]) & res.iloc[:,2].isin(templates[1].iloc[:,2])].drop_duplicates()
            wrong_Tier2['Wrong Tier'] ='Tier 2'
            wrong_Tier2['Possible match Tier 1']=np.nan
            wrong_Tier2['Possible match Tier 2']=wrong_Tier2.iloc[:,3].apply(lambda x: x if pd.isnull(x) or type(x)==float or type(x)==int else get_close_matches(x,templates[1].iloc[:,3].astype(str).unique().tolist(),1)).apply(lambda x: 'Not Found' if len(x)==0 or pd.isnull(x) or type(x)==float or type(x)==int else ''.join(x))
            wrong_Tier1_2=res.loc[~res.iloc[:,2].isin(templates[1].iloc[:,2]) & ~res.iloc[:,3].isin(templates[1].iloc[:,3])].drop_duplicates()
            wrong_Tier1_2['Wrong Tier'] ='Tier 1 and 2'
            wrong_Tier1_2['Possible match Tier 1']=wrong_Tier1_2.iloc[:,2].apply(lambda x: x if pd.isnull(x) or type(x)==float or type(x)==int else get_close_matches(x,templates[1].iloc[:,2].astype(str).unique().tolist(),1)).apply(lambda x: 'Not Found' if len(x)==0 or pd.isnull(x) or type(x)==float or type(x)==int else ''.join(x))
            wrong_Tier1_2['Possible match Tier 2']=wrong_Tier1_2.iloc[:,3].apply(lambda x: x if pd.isnull(x) or type(x)==float or type(x)==int else get_close_matches(x,templates[1].iloc[:,3].astype(str).unique().tolist(),1)).apply(lambda x: 'Not Found' if len(x)==0 or pd.isnull(x) or type(x)==float or type(x)==int else ''.join(x))
            wrong_Tiers=pd.concat([wrong_Tier1,wrong_Tier2,wrong_Tier1_2],axis=0)
            wrong_Tiers.dropna(axis='columns',how='all',inplace=True)
            #wrong_res=wrong_Tiers.iloc[:,[0,1,2,3,4,7,5,6]]
            with pd.ExcelWriter(op_res_cats_report + company + '_res_cats_issues_'+ dt.datetime.now().strftime("%Y-%m-%d %H-%M-%S") +'.xlsx',engine='xlsxwriter') as writer:
                res.to_excel(writer, 'ResCats',index=False)
                if len(wrong_Tiers)>0:
                    print('','COUNT: '+str(np.shape(wrong_Tiers)[0]),'',sep='\n',file=open(op_res_cats_report +'res_issues.txt','a',encoding='utf8'))
                    wrong_Tiers.to_excel(writer, 'wrong_res',index=False)
                else:
                    print('',sep='\n',file=open(op_res_cats_report +'res_issues.txt','a',encoding='utf8'))
                writer.save()
        else:
            None
        if len(ops)>0:
            wrong_Tier1=ops.loc[~ops.iloc[:,2].isin(templates[0].iloc[:,2]) & ops.iloc[:,3].isin(templates[0].iloc[:,3])].drop_duplicates()
            wrong_Tier1['Wrong Tier'] ='Tier 1'
            wrong_Tier1['Possible match Tier 1']=wrong_Tier1.iloc[:,2].apply(lambda x: x if pd.isnull(x) or type(x)==float or type(x)==int else get_close_matches(x,templates[0].iloc[:,2].astype(str).unique().tolist(),1)).apply(lambda x: 'Not Found' if len(x)==0 or pd.isnull(x) or type(x)==float or type(x)==int else ''.join(x))
            wrong_Tier1['Possible match Tier 2']=np.nan
            wrong_Tier2=ops.loc[~ops.iloc[:,3].isin(templates[0].iloc[:,3]) & ops.iloc[:,2].isin(templates[0].iloc[:,2])].drop_duplicates()
            wrong_Tier2['Wrong Tier'] ='Tier 2'
            wrong_Tier2['Possible match Tier 1']=np.nan
            wrong_Tier2['Possible match Tier 2']=wrong_Tier2.iloc[:,3].apply(lambda x: x if pd.isnull(x) or type(x)==float or type(x)==int else get_close_matches(x,templates[0].iloc[:,3].astype(str).unique().tolist(),1)).apply(lambda x: 'Not Found' if len(x)==0 or pd.isnull(x) or type(x)==float or type(x)==int else ''.join(x))
            wrong_Tier1_2=ops.loc[~sheets[j].iloc[:,2].isin(templates[0].iloc[:,2]) & ~ops.iloc[:,3].isin(templates[0].iloc[:,3])].drop_duplicates()
            wrong_Tier1_2['Wrong Tier'] ='Tier 1 and 2'
            wrong_Tier1_2['Possible match Tier 1']=wrong_Tier1_2.iloc[:,2].apply(lambda x: x if pd.isnull(x) or type(x)==float or type(x)==int else get_close_matches(x,templates[0].iloc[:,2].astype(str).unique().tolist(),1)).apply(lambda x: 'Not Found' if len(x)==0 or pd.isnull(x) or type(x)==float or type(x)==int else ''.join(x))
            wrong_Tier1_2['Possible match Tier 2']=wrong_Tier1_2.iloc[:,3].apply(lambda x: x if pd.isnull(x) or type(x)==float or type(x)==int else get_close_matches(x,templates[0].iloc[:,3].astype(str).unique().tolist(),1)).apply(lambda x: 'Not Found' if len(x)==0 or pd.isnull(x) or type(x)==float or type(x)==int else ''.join(x))
            wrong_Tiers=pd.concat([wrong_Tier1,wrong_Tier2,wrong_Tier1_2],axis=0)
            wrong_Tiers.dropna(axis='columns',how='all',inplace=True)
            #wrong_ops=wrong_Tiers.iloc[:,[0,1,2,3,4,7,5,6]]
            with pd.ExcelWriter(op_res_cats_report + company + '_op_cats_issues_'+ dt.datetime.now().strftime("%Y-%m-%d %H-%M-%S") +'.xlsx',engine='xlsxwriter') as writer:
                ops.to_excel(writer, 'wrong_ops',index=False)
                if len(wrong_Tiers)>0:
                    print('','COUNT: '+str(np.shape(wrong_Tiers)[0]),'',sep='\n',file=open(op_res_cats_report +'op_issues.txt','a',encoding='utf8'))
                    wrong_Tiers.to_excel(writer, 'wrong_ops',index=False)
                else:
                    print('',sep='\n',file=open(op_res_cats_report +'op_issues.txt','a',encoding='utf8'))
                writer.save()
        else:
            None    