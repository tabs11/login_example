# -*- coding: utf-8 -*-
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import datetime as dt
import re
import glob
from difflib import get_close_matches
import itertools
from tabulate import tabulate

def op_res_cats_files(path,company,op_res_cats_report):
    #print('','#'*47,'#' +' Operational Resolution Categories validation'.upper()+ '#','#'*47,'',sep='\n',file=open(op_res_cats_report +'issues.txt','a',encoding='utf8'))
    ##Operational Category
    op_cats=load_workbook('CMDB_templates/OperationalCatalog.xlsm',read_only=False, keep_vba=True)
    sheets_ops = op_cats.sheetnames
    w_sheet_op_cats=op_cats[sheets_ops[1]]
    w_sheet1_op_cats=op_cats[sheets_ops[2]]
    #Resolution Category
    res_cats=load_workbook('CMDB_templates/Generic_Catalog.xlsm',read_only=False, keep_vba=True)
    sheets_res = res_cats.sheetnames
    w_sheet1_res_cats=res_cats[sheets_res[1]]
    w_sheet2_res_cats=res_cats[sheets_res[2]]
    w_sheet3_res_cats=res_cats[sheets_res[3]]
    sheets=[]
    res=pd.DataFrame()
    ops=pd.DataFrame()
    templates=[]
    op_template=pd.read_excel('Op_Res_Cats/Datamodel OpCats-ResCats-LoV.xlsx',pd.ExcelFile('Op_Res_Cats/Datamodel OpCats-ResCats-LoV.xlsx').sheet_names[0])
    res_template=pd.read_excel('Op_Res_Cats/Datamodel OpCats-ResCats-LoV.xlsx',pd.ExcelFile('Op_Res_Cats/Datamodel OpCats-ResCats-LoV.xlsx').sheet_names[1])
    templates.append(op_template)
    templates.append(res_template)
    op_col=op_template.columns.tolist()
    res_col=res_template.columns.tolist()
    ####check colnames and sheetnames
    sheetnames=['RES.CAT.','RES CAT','ResCat','OP.CAT.','OP CAT','OpCat']
    unmatched_fields_res=[]
    unmatched_fields_ops=[]
    for j in range(len(glob.glob(path+'/*'))):
        if glob.glob(path+'/*')[j].endswith(('.xls','.xlsx')):
            for k in range(len(list(set(pd.ExcelFile(glob.glob(path+'*')[j]).sheet_names).intersection(sheetnames)))):
                sheets.append(pd.read_excel(glob.glob(path+'*')[j],list(set(pd.ExcelFile(glob.glob(path+'*')[j]).sheet_names).intersection(sheetnames))[k]))
            #files=pd.read_excel(glob.glob(path+'*')[j],sheet_name=None)
            #for frame in files.keys():
            #   sheets.append(files[frame])
        elif glob.glob(path+'/*')[j].endswith('.csv'):
            sheets.append(pd.read_csv(glob.glob(path+'*')[j],sep=";",encoding='ISO-8859-1'))
        else:
            None
    #if len(sheets)==0:
    #    print('','Check the sheetnames in the input files','The values should be: ResCat or/and OpCat',sep='\n',file=open(op_res_cats_report +'Mismatched_fields.txt','a',encoding='utf8'))
    #else:
        ###global checks
    for j in range(len(sheets)):
        sheets[j].fillna('',inplace=True)
        for k in range(len(sheets[j].columns)):
            sheets[j].iloc[:,k]=pd.Series(np.where(sheets[j].iloc[:,k].str.contains('None'),'- None -',sheets[j].iloc[:,k]))
            sheets[j].iloc[:,k]=sheets[j].iloc[:,k].apply(lambda x: x.strip() if type(x)==str else x)
        #res
        if (sheets[j].columns.str.contains('ResCat',case=False).any()):
            res=sheets[j]
            unmatched_fields_res.append(list(set(res_col) - set(res)))
        else:
            None
        #ops
        if sheets[j].columns.str.contains('OpCat',case=False).any():
            ops=sheets[j]
            unmatched_fields_ops.append(list(set(op_col) - set(ops)))
        else:
            None
        sheets[j]['Company']=company
    if len(res)>0:
        if len(list(itertools.chain(*unmatched_fields_res)))>0:
            unmatched=pd.DataFrame(pd.Series(list(itertools.chain(*unmatched_fields_res))).rename('FIELD'))
            print('','Resolution Categories template issue:','-'*len('Resolution Categories template issue:'),tabulate(unmatched,headers="keys",tablefmt="fancy_grid",showindex=False),'',sep='\n',file=open(op_res_cats_report +'Mismatched_fields_res.txt','a',encoding='utf8'))
        else:
            
            wrong_Tier1=res.loc[~res.iloc[:,2].isin(templates[1].iloc[:,2]) & res.iloc[:,3].isin(templates[1].iloc[:,3])].drop_duplicates()
            wrong_Tier1['Wrong Tier'] ='Tier 1'
            wrong_Tier1['Possible match Tier 1']=wrong_Tier1.iloc[:,2].apply(lambda x: x if pd.isnull(x) or type(x)==float or type(x)==int else get_close_matches(x,templates[1].iloc[:,2].astype(str).unique().tolist(),1)).apply(lambda x: 'Not Found' if len(x)==0 or pd.isnull(x) or type(x)==float or type(x)==int else ''.join(x))
            wrong_Tier1['Possible match Tier 2']=np.nan
            wrong_Tier2=res.loc[~res.iloc[:,3].isin(templates[1].iloc[:,3]) & res.iloc[:,2].isin(templates[1].iloc[:,2])].drop_duplicates()
            wrong_Tier2['Wrong Tier'] ='Tier 2'
            wrong_Tier2['Possible match Tier 1']=np.nan
            wrong_Tier2['Possible match Tier 2']=wrong_Tier2.iloc[:,3].apply(lambda x: x if pd.isnull(x) or type(x)==float or type(x)==int else get_close_matches(x,templates[1].iloc[:,3].astype(str).unique().tolist(),1)).apply(lambda x: 'Not Found' if len(x)==0 or pd.isnull(x) or type(x)==float or type(x)==int else ''.join(x))
            wrong_Tier1_2=res.loc[~res.iloc[:,2].isin(templates[1].iloc[:,2]) & ~res.iloc[:,3].isin(templates[1].iloc[:,3])].drop_duplicates()
            wrong_Tier1_2['Wrong Tier'] ='Tier 1 and 2'
            wrong_Tier1_2['Possible match Tier 1']=wrong_Tier1_2.iloc[:,2].apply(lambda x: x if pd.isnull(x) or type(x)==float or type(x)==int else get_close_matches(x,templates[1].iloc[:,2].astype(str).unique().tolist(),1)).apply(lambda x: 'Not Found' if len(x)==0 or pd.isnull(x) or type(x)==float or type(x)==int else ''.join(x))
            wrong_Tier1_2['Possible match Tier 2']=wrong_Tier1_2.iloc[:,3].apply(lambda x: x if pd.isnull(x) or type(x)==float or type(x)==int else get_close_matches(x,templates[1].iloc[:,3].astype(str).unique().tolist(),1)).apply(lambda x: 'Not Found' if len(x)==0 or pd.isnull(x) or type(x)==float or type(x)==int else ''.join(x))
            wrong_Tiers=pd.concat([wrong_Tier1,wrong_Tier2,wrong_Tier1_2],axis=0)
            wrong_Tiers.dropna(axis='columns',how='all',inplace=True)
            #wrong_res=wrong_Tiers.iloc[:,[0,1,2,3,4,7,5,6]]
            if len(wrong_Tiers)>0:
                print('',tabulate(wrong_Tiers,headers="keys",tablefmt="fancy_grid",showindex=False),'',sep='\n',file=open(op_res_cats_report +'res_issues.txt','a',encoding='utf8'))

                res.reset_index(inplace=True)
                wrong_Tiers.reset_index(inplace=True)
                wrong_Tiers=res.merge(wrong_Tiers.drop(columns=['Company','Module','ResCat1','ResCat2','ResCat3']),on='index', how='outer')
                with pd.ExcelWriter(op_res_cats_report + company + '_res_cats_issues'+ dt.datetime.now().strftime("%Y-%m-%d %H-%M-%S") +'.xlsx',engine='xlsxwriter') as writer:
                    wrong_Tiers.to_excel(writer, 'wrong_res',index=False)
                    writer.save()

            else:
                print('',sep='\n',file=open(op_res_cats_report +'res_issues.txt','a',encoding='utf8'))
            for i in range(res.shape[0]):
                w_sheet1_res_cats['A' +str(4+i)]='Resolution Category'
                w_sheet1_res_cats['B' +str(4+i)]=res.filter(regex=re.compile('ResCat1',re.IGNORECASE)).iloc[:,0].values[i]
                w_sheet1_res_cats['C' +str(4+i)]=res.filter(regex=re.compile('ResCat2',re.IGNORECASE)).iloc[:,0].values[i]
                w_sheet1_res_cats['D' +str(4+i)]=res.filter(regex=re.compile('ResCat3',re.IGNORECASE)).iloc[:,0].values[i]
                w_sheet1_res_cats['O' +str(4+i)]='Enabled'
                w_sheet2_res_cats['A' +str(4+i)]=res.filter(regex=re.compile('Company',re.IGNORECASE)).iloc[:,0].values[i]
                w_sheet2_res_cats['B' +str(4+i)]='Enabled'
                w_sheet2_res_cats['C' +str(4+i)]='Resolution Category'
                w_sheet2_res_cats['D' +str(4+i)]=res.filter(regex=re.compile('ResCat1',re.IGNORECASE)).iloc[:,0].values[i]
                w_sheet2_res_cats['E' +str(4+i)]=res.filter(regex=re.compile('ResCat2',re.IGNORECASE)).iloc[:,0].values[i]
                w_sheet2_res_cats['F' +str(4+i)]=res.filter(regex=re.compile('ResCat3',re.IGNORECASE)).iloc[:,0].values[i]
                w_sheet2_res_cats['L' +str(4+i)]='Yes'
                w_sheet2_res_cats['M' +str(4+i)]='Yes'
                w_sheet2_res_cats['N' +str(4+i)]='Yes'
                w_sheet2_res_cats['O' +str(4+i)]='Yes'
                w_sheet2_res_cats['P' +str(4+i)]='Yes'
                w_sheet3_res_cats['A' +str(4+i)]='Resolution Category'
                w_sheet3_res_cats['B' +str(4+i)]=res.filter(regex=re.compile('ResCat1',re.IGNORECASE)).iloc[:,0].values[i]
                w_sheet3_res_cats['C' +str(4+i)]=res.filter(regex=re.compile('ResCat2',re.IGNORECASE)).iloc[:,0].values[i]
                w_sheet3_res_cats['D' +str(4+i)]=res.filter(regex=re.compile('ResCat3',re.IGNORECASE)).iloc[:,0].values[i]
                w_sheet3_res_cats['F' +str(4+i)]='Enabled'
            res_cats.save(filename = op_res_cats_report + company + '_res_cats_NOAM_' + dt.datetime.now().strftime("%Y-%m-%d %H-%M-%S") + '.xlsm')
    else:
        None
    if len(ops)>0:
        if len(list(itertools.chain(*unmatched_fields_ops)))>0:
            unmatched=pd.DataFrame(pd.Series(list(itertools.chain(*unmatched_fields_ops))).rename('FIELD'))
            print('','Opeational Categories template issue:','-'*len('Opeational Categories template issue:'),tabulate(unmatched,headers="keys",tablefmt="fancy_grid",showindex=False),'',sep='\n',file=open(op_res_cats_report +'Mismatched_fields_ops.txt','a',encoding='utf8'))
        else:
            wrong_Tier1=ops.loc[~ops.iloc[:,2].isin(templates[0].iloc[:,2]) & ops.iloc[:,3].isin(templates[0].iloc[:,3])].drop_duplicates()
            wrong_Tier1['Wrong Tier'] ='Tier 1'
            wrong_Tier1['Possible match Tier 1']=wrong_Tier1.iloc[:,2].apply(lambda x: x if pd.isnull(x) or type(x)==float or type(x)==int else get_close_matches(x,templates[0].iloc[:,2].astype(str).unique().tolist(),1)).apply(lambda x: 'Not Found' if len(x)==0 or pd.isnull(x) or type(x)==float or type(x)==int else ''.join(x))
            wrong_Tier1['Possible match Tier 2']=np.nan
            wrong_Tier2=ops.loc[~ops.iloc[:,3].isin(templates[0].iloc[:,3]) & ops.iloc[:,2].isin(templates[0].iloc[:,2])].drop_duplicates()
            wrong_Tier2['Wrong Tier'] ='Tier 2'
            wrong_Tier2['Possible match Tier 1']=np.nan
            wrong_Tier2['Possible match Tier 2']=wrong_Tier2.iloc[:,3].apply(lambda x: x if pd.isnull(x) or type(x)==float or type(x)==int else get_close_matches(x,templates[0].iloc[:,3].astype(str).unique().tolist(),1)).apply(lambda x: 'Not Found' if len(x)==0 or pd.isnull(x) or type(x)==float or type(x)==int else ''.join(x))
            wrong_Tier1_2=ops.loc[~ops.iloc[:,2].isin(templates[0].iloc[:,2]) & ~ops.iloc[:,3].isin(templates[0].iloc[:,3])].drop_duplicates()
            wrong_Tier1_2['Wrong Tier'] ='Tier 1 and 2'
            wrong_Tier1_2['Possible match Tier 1']=wrong_Tier1_2.iloc[:,2].apply(lambda x: x if pd.isnull(x) or type(x)==float or type(x)==int else get_close_matches(x,templates[0].iloc[:,2].astype(str).unique().tolist(),1)).apply(lambda x: 'Not Found' if len(x)==0 or pd.isnull(x) or type(x)==float or type(x)==int else ''.join(x))
            wrong_Tier1_2['Possible match Tier 2']=wrong_Tier1_2.iloc[:,3].apply(lambda x: x if pd.isnull(x) or type(x)==float or type(x)==int else get_close_matches(x,templates[0].iloc[:,3].astype(str).unique().tolist(),1)).apply(lambda x: 'Not Found' if len(x)==0 or pd.isnull(x) or type(x)==float or type(x)==int else ''.join(x))
            wrong_Tiers=pd.concat([wrong_Tier1,wrong_Tier2,wrong_Tier1_2],axis=0)
            wrong_Tiers.dropna(axis='columns',how='all',inplace=True)
            #wrong_ops=wrong_Tiers.iloc[:,[0,1,2,3,4,7,5,6]]
            if len(wrong_Tiers)>0:
                print('',tabulate(wrong_Tiers,headers="keys",tablefmt="fancy_grid",showindex=False),'',sep='\n',file=open(op_res_cats_report +'op_issues.txt','a',encoding='utf8'))
                ops.reset_index(inplace=True)
                wrong_Tiers.reset_index(inplace=True)
                wrong_Tiers=ops.merge(wrong_Tiers.drop(columns=['Company','Module','OpCat1','OpCat2','OpCat3']),on='index', how='outer')
                with pd.ExcelWriter(op_res_cats_report + company + '_op_cats_issues'+ dt.datetime.now().strftime("%Y-%m-%d %H-%M-%S") +'.xlsx',engine='xlsxwriter') as writer:
                    wrong_Tiers.to_excel(writer, 'wrong_ops',index=False)
                    writer.save()
            else:
                print('',sep='\n',file=open(op_res_cats_report +'op_issues.txt','a',encoding='utf8'))
            ops['inc_values'] = np.where(ops['Module'] == "Incident Management", 'Yes', None)
            ops['prb_values'] = np.where(ops['Module'] == "Problem Management", 'Yes', None)
            ops['chg_values'] = np.where(ops['Module'] == "Change Management", 'Yes', None)
            ops['conf_values'] = np.where(ops['Module'] == "Configuration/Asset Management", 'Yes', None)
            ops['rel_values'] = np.where(ops['Module'] == "Release Management", 'Yes', None)
            for i in range(ops.shape[0]): 
                w_sheet_op_cats['A' +str(4+i)]=ops.filter(regex=re.compile('Company',re.IGNORECASE)).iloc[:,0].values[i]
                w_sheet_op_cats['B' +str(4+i)]=ops.filter(regex=re.compile('OpCat1',re.IGNORECASE)).iloc[:,0].values[i]
                w_sheet_op_cats['C' +str(4+i)]=ops.filter(regex=re.compile('OpCat2',re.IGNORECASE)).iloc[:,0].values[i]
                w_sheet_op_cats['D' +str(4+i)]=ops.filter(regex=re.compile('OpCat3',re.IGNORECASE)).iloc[:,0].values[i]
                w_sheet_op_cats['Q' +str(4+i)]='Enabled'
                w_sheet1_op_cats['A' +str(4+i)]=ops.filter(regex=re.compile('OpCat1',re.IGNORECASE)).iloc[:,0].values[i]
                w_sheet1_op_cats['B' +str(4+i)]=ops.filter(regex=re.compile('OpCat2',re.IGNORECASE)).iloc[:,0].values[i]
                w_sheet1_op_cats['C' +str(4+i)]=ops.filter(regex=re.compile('OpCat3',re.IGNORECASE)).iloc[:,0].values[i]
                w_sheet1_op_cats['F' +str(4+i)]='Enabled'
                w_sheet_op_cats['F' +str(4+i)]=ops['inc_values'].values[i]
                w_sheet_op_cats['G' +str(4+i)]=ops['inc_values'].values[i]
                w_sheet_op_cats['H' +str(4+i)]=ops['inc_values'].values[i]
                w_sheet_op_cats['I' +str(4+i)]=ops['inc_values'].values[i]
                w_sheet_op_cats['J' +str(4+i)]=ops['inc_values'].values[i]
                w_sheet_op_cats['K' +str(4+i)]=ops['prb_values'].values[i]
                w_sheet_op_cats['N' +str(4+i)]=ops['chg_values'].values[i]
                w_sheet_op_cats['L' +str(4+i)]=ops['conf_values'].values[i]
                w_sheet_op_cats['R' +str(4+i)]=ops['rel_values'].values[i]
            op_cats.save(filename = op_res_cats_report + company + '_op_cats_NOAM_' + dt.datetime.now().strftime("%Y-%m-%d %H-%M-%S") + '.xlsm')
    else:
        None