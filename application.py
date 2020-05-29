#!/usr/bin/env python
from lib import *
from flask import Flask, request, redirect, url_for, render_template, send_from_directory,send_file,flash,session
from werkzeug.utils import secure_filename
from flask_login import LoginManager
from flask_login import UserMixin # subclass of flask user
from flask_login import login_required
from flask_login import login_user
from flask_login import logout_user
import process_data
import process_noam_data
import process_res_cats
import process_res_cats_test
import process_zte
import update_priority as up_prio
#import process_site_history
import process_site_history_users
import process_cmdb_inventory
import process_rap
import process_rap_noam
import process_prod_cat
import process_cmdb_update
from openpyxl import load_workbook
#import flask
#import flask_saml
#import process_common_val
#import process_specific_val


class User(UserMixin):
	def __init__(self, username,password):
		super(User, self).__init__()
		self.username = username
		self.password = password
	def is_active(self):
		return True
	def is_anonymous(self):
		return False
	def get_id(self):
		return self.username

# USER DATABASE
USERS = { # dictionary (username, User)
	'numartin' : User('numartin','pass'),
	'jschuur' : User('jschuur','teste'),
	'nawaz' : User('nawaz','nawaz'),
	'mccavitt' : User('mccavitt','mccavitt'),
	'abociat' : User('abociat','abociat'),
	'mariusgc' : User('mariusgc','mariusgc'),
	'gelei' : User('gelei','gelei'),
	'marcinkl' : User('marcinkl','marcinkl'),
	'mayayada' : User('mayayada','mayayada'),
	'dees' : User('dees','dees'),
	'singaram' : User('singaram','singaram'),
	'nvp' : User('nvp','nvp'),
	'avilcean' : User('avilcean','avilcean'),
	'paulof'   : User('paulof','paulof'),
	'adriani'  : User('adriani','adriani'),
	'paagrawa'  : User('paagrawa','paagrawa'),
	'paagrawa_admin'  : User('paagrawa_admin','paagrawa_admin'),
	'mlacan'  : User('mlacan','mlacan'),
	'rserban' : User('rserban','rserban'),
	'matheka' : User('matheka','matheka'),
	'fjoseph' : User('fjoseph','fjoseph'),
	'difrance' : User('difrance','difrance'),
	'canuto'   : User('canuto','canuto'),
	'ciszewsk' : User('ciszewsk','ciszewsk'),
	'sancosta' : User('sancosta','sancosta'),
	'tdopiera' : User('tdopiera','tdopiera'),
	'hassane'  : User('hassane','hassane'),
	'sfikisz'  : User('sfikisz','sfikisz'),
	'emomal'   : User('emomal','emomal'),
	'jsantana' : User('jsantana','jsantana'),
	'ssoaresd' : User('ssoaresd','ssoaresd'),
	'jzeglicz' : User('jzeglicz','jzeglicz'),
	'wmok':User('wmok','wmok'),
	'gocs':User('gocs','gocs'),
	'camaelo':User('camaelo','camaelo'),
	'cmirceai':User('cmirceai','cmirceai'),
	'duncant':User('duncant','duncant'),
	'roprita':User('roprita','roprita'),
	'copene':User('copene','copene'),
	'mtn_sig_mayayada':User('mtn_sig_mayayada','mtn_sig_mayayada'),
	'arunrm':User('arunrm','arunrm'),
	'kramanuj':User('kramanuj','kramanuj'),
	'pedrorod':User('pedrorod','pedrorod'),
	'pt100712':User('pt100712','pt100712'),
	'andrea':User('andrea','andrea'),
	'ashmarh':User('ashmarh','ashmarh'),
	'ancam':User('ancam','ancam'),
	'ancam':User('ancam','ancam'),
	'arunrm':User('arunrm','arunrm'),
	'abhanand':User('abhanand','abhanand'),
	'rathod':User('rathod','rathod'),
	'roprita':User('roprita','roprita'),
	'bharam':User('bharam','bharam'),
	'arotaru':User('arotaru','arotaru'),
	'bnanu':User('bnanu','bnanu')

		
}

application = Flask(__name__)
#application.config.update({
##    'SECRET_KEY': 'soverysecret',
#    'SAML_METADATA_URL': 'https://login.microsoftonline.com/5d471751-9675-428d-917b-70f44f9630b0/federationmetadata/2007-06/federationmetadata.xml?appid=21c725a9-b27e-4221-9864-adb4da9edfa4',
#})
#flask_saml.FlaskSAML(application)
#
SECRET_KEY='bla'
application.secret_key = SECRET_KEY
#CMDB_FOLDER = 'CMDB_templates/'
#application.config['CMDB_FOLDER']=CMDB_FOLDER
application.config['ALLOWED_EXTENSIONS'] = set(['xlsx','xls','csv'])





# login views
@application.route('/login', methods=['GET'])
def login_get():

	return render_template('login.html')

@application.route('/login', methods=['POST'])
def login_post():

	# get details from post request
	username = request.form['username']
	password = request.form['password']
	# get  user
	try:
		user = USERS[username]
		session['username']=str(username)
	except KeyError:
		user = None
	# validate user
	if user and user.password == password:
		login_user(user)
		if request.args.get("next"):
			return redirect(request.args.get("next"))
		else:
			return redirect('/')
	else:
		flash("Wrong Username or Password")
		return render_template('login.html')



def allowed_file(filename):
	return '.' in filename and \
		   filename.rsplit('.', 1)[1] in application.config['ALLOWED_EXTENSIONS']



@application.route('/user_guide', methods=['GET'])
def user_guide():

	return render_template('user_guide.html')

@application.route('/user_guide_v2', methods=['GET'])
def user_guide_v2():

	return render_template('user_guide_v2.html')

##########################################################################
#####download templates
#@application.route('/return-file/')
#@login_required
#def return_file():
#	filename='Templates.zip'
#	return send_file(os.path.join(application.config['CMDB_FOLDER'])+filename,attachment_filename=filename, as_attachment=True)

#@application.route('/templates')
#@login_required
#def file_downloads():
#	return render_template('home.html')


@application.route('/action', methods=['GET'])
def index():
	return redirect("/", code=302)





###########################################################################
@application.route('/')
@login_required
def drop():
	data=pd.DataFrame()
	user=session['username']
	cmdb_owners=pd.read_excel('CMDB_templates/cmdb Owners_full_list.xlsx')
	if user in ['numartin','paulof','mccavitt','paagrawa','hassane','bnanu']:
		data=cmdb_owners['Company']
	else:
		data=cmdb_owners['Company'][cmdb_owners['Login ID']==user]
		
	data=np.unique(data).tolist()
	data.insert(0,'Dummy Company')
	return render_template(
		'home.html',data=data
		)



@application.route("/action" , methods=['GET', 'POST'])
@login_required
def home():

	data=[s for s in os.listdir(os.getcwd()) if len(s) > 30]
	paths_to_del=[]
	dates=[]
	for i in range(len(data)):
		paths_to_del.append(os.getcwd()+ '/' + data[i])
		dates.append((dt.datetime.now()-datetime.fromtimestamp(os.path.getctime(paths_to_del[i]))).days)
		if dates[i]>0:
			shutil.rmtree(paths_to_del[i])	
		else:
			None
	select = request.form.get('comp_select')
	if request.method == 'POST':
		user=session['username']
		session['company']=str(select)
		#session['filename']=session['company']+'_'+str(uuid.uuid1())
		#ID_FOLDER=session['filename']
		#UPLOAD_FOLDER=ID_FOLDER + '/Files_to_validate/'
		#os.makedirs(ID_FOLDER)
		#os.makedirs(UPLOAD_FOLDER)
		msg=session['company']
			#msg = 'Successfull'
		if user!='numartin' and user!='paulof':
			#return render_template('multi_upload_index.html')
			return render_template('cmdb_validation_users.html',msg=msg)
		else:
			return render_template('validation_conversion_admin.html',msg=msg)


@application.route("/validation" , methods=['GET', 'POST'])
@login_required
def validation():
	msg=session['company']
	return render_template('cmdb_validation_admin.html',msg=msg)

@application.route("/conversion" , methods=['GET', 'POST'])
@login_required
def conversion():
	msg=session['company']
	return render_template('cmdb_conversion_admin.html',msg=msg)



#######CMDB reports#############
@application.route('/site_upload', methods=['GET'])
@login_required
def site_upload():
	msg_cmdb1=None
	msg_cmdb2=None
	user=session['username']
	session['filename']=session['company']+'_'+str(uuid.uuid1())
	ID_FOLDER=session['filename']
	DOWNLOAD_FOLDER=ID_FOLDER +'/Report/'
	os.makedirs(ID_FOLDER)
	os.makedirs(DOWNLOAD_FOLDER)
	msg_company=ID_FOLDER.split('_')[0]
	#process_site_history.sites_cis_report(company=ID_FOLDER.split('_')[0],site_report=DOWNLOAD_FOLDER)
	process_site_history_users.sites_cis_report(user=user,company=ID_FOLDER.split('_')[0],site_report=DOWNLOAD_FOLDER)
	site_filenames=[f for f in os.listdir(DOWNLOAD_FOLDER) if f.endswith(('.xlsx','csv'))]
	if 'SQLDB_CMDB.txt' in os.listdir(DOWNLOAD_FOLDER):
		text_cmdb=open(DOWNLOAD_FOLDER+'SQLDB_CMDB.txt', 'r+',encoding='utf8')
		content_cmdb = text_cmdb.read()
		text_cmdb.close()
		if 'found in inven' in content_cmdb:
			msg_cmdb2='Missing CMDB inventory'.upper()
		else:
			msg_cmdb1=ID_FOLDER.split('_')[0].upper() + ' Current CMDB size in ITSM: '.upper()
	else:
		msg_cmdb2='Missing CMDB inventory'.upper()
	return render_template('site_upload.html', 
		site_filenames=site_filenames,
		msg_company=msg_company,
		msg_cmdb1=msg_cmdb1,
		text_cmdb=content_cmdb,
		msg_cmdb2=msg_cmdb2)


@application.route('/site_report/<filename>')
@login_required
def uploaded_site_file(filename):
	ID_FOLDER=session['filename']
	DOWNLOAD_FOLDER=ID_FOLDER +'/Report/'
	return send_from_directory(DOWNLOAD_FOLDER,filename)


####update prodCat############
@application.route('/update_prod_cat', methods=['GET','POST'])
@login_required
def update_prod_cat():
	msg_company=None
	msg3=None
	session['filename']=session['company']+'_'+str(uuid.uuid1())
	ID_FOLDER=session['filename']
	#TEMP_FOLDER=ID_FOLDER +'/new_Prod_Cats/'
	UPLOAD_FOLDER='Prod_Cats_V2/'

	#shutil.rmtree('Prod_Cats_V2/')	
	uploaded_files = request.files.getlist("file[]")
	for file in uploaded_files:
		# Check if the file is one of the allowed types/extensions
		if file and allowed_file(file.filename):
			
			# Make the filename safe, remove unsupported chars
			filename = secure_filename(file.filename)
			if not os.path.exists(UPLOAD_FOLDER):
				os.makedirs(UPLOAD_FOLDER)
			# Move the file form the temporal folder to the upload
			file.save(os.path.join(UPLOAD_FOLDER, filename))
			filenames=os.listdir(UPLOAD_FOLDER)
			msg3=filenames
		else:
			msg3='Please select a valid extension (.xls(x) or .csv)'

	return render_template('update_prod_cat.html',msg3=msg3)

#######upload prod_cat

#@application.route('/prod_upload', methods=['GET'])
#@login_required
#def prod_upload():
#	ID_FOLDER=session['filename']
#	TEMP_FOLDER=ID_FOLDER +'/new_Prod_Cats/'
#	DOWNLOAD_FOLDER= 'Prod_Cats_V2/'
#	process_prod_cat.prod_update(path=TEMP_FOLDER)
#	prod_filenames=[f for f in os.listdir(DOWNLOAD_FOLDER) if f.endswith(('.xlsx','csv'))]
#	return render_template('prod_update.html',prod_filenames=prod_filenames)
#
#
#
#@application.route('/report_prod/<filename>')
#@login_required
#def uploaded_prod(filename):
#	DOWNLOAD_FOLDER= 'Prod_Cats_V2/'
#	return send_from_directory(DOWNLOAD_FOLDER,filename)
######################################################################



##########################################################################
#####download prod_cat
@application.route('/return-file/')
#@login_required
def return_file():
	filename='oneitsm_ProdCats.csv'
	return send_file('Prod_Cats_V2/'+filename,attachment_filename=filename, as_attachment=True)

@application.route('/templates')
@login_required
def file_downloads():
	return render_template('home.html')


#####input files to validate

@application.route('/data', methods=['GET','POST'])
@login_required
def data_to_validate():
	urls=None
	user=session['username']
	company=session['company']
	cmdb_owners=pd.read_excel('CMDB_templates/cmdb Owners_full_list.xlsx')
	if company=='Dummy Company':
		urls=''
		email=''
		name=''
	else:
		links=cmdb_owners['Data Model'][((cmdb_owners['Login ID']==user) | (user in ['numartin','paulof','mccavitt','paagrawa'])) & (cmdb_owners['Company']==company)]
		name=list(cmdb_owners['First Name*'][cmdb_owners['Company']==company] + ' '+cmdb_owners['Last Name*+'][cmdb_owners['Company']==company])
		email=list(cmdb_owners['Email Address'][cmdb_owners['Company']==company])
		if (links=='No hyperlink').any():
			urls=''
		else: 
			urls=str(np.unique(links).tolist()[0])
	session['urls']=urls
	session['name']=name
	session['email']=email
	msg_company=None
	msg3=None
	msg_to_many_files=None
	session['filename']=session['company']+'_'+str(uuid.uuid1())
	ID_FOLDER=session['filename']
	msg_company=ID_FOLDER.split('_')[0]
	UPLOAD_FOLDER=ID_FOLDER + '/Files_to_validate/'
	DOWNLOAD_FOLDER=ID_FOLDER+'/Report/'
	uploaded_files = request.files.getlist("file[]")
	for file in uploaded_files:
		# Check if the file is one of the allowed types/extensions
		if file and allowed_file(file.filename):
			
			# Make the filename safe, remove unsupported chars
			filename = secure_filename(file.filename)
			
			if not os.path.exists(ID_FOLDER):
				os.makedirs(ID_FOLDER)
				os.makedirs(UPLOAD_FOLDER)
				os.makedirs(DOWNLOAD_FOLDER)
			# Move the file form the temporal folder to the upload
			file.save(os.path.join(UPLOAD_FOLDER, filename))
			filenames=os.listdir(UPLOAD_FOLDER)
			if len(os.listdir(UPLOAD_FOLDER))>2:
				msg_to_many_files='THE NUMBER OF INPUT FILES SHOULD BE 1 (sites or cis) OR 2(sites and cis). PLEASE SELECT AGAIN THE INPUT FILES.'

			else:
				msg3='File(s) Successfully Submited. Click ont the button bellow to start the validation.'
		else:
			msg3='Please select a valid extension (.xls or .xlsx)'

	return render_template('multi_upload_index.html',msg3=msg3,msg_company=msg_company,urls=urls,name=name,email=email,msg_to_many_files=msg_to_many_files)

@application.route('/upload', methods=['POST'])
#@cache.cached(timeout=500)
@login_required
def upload():
	urls=session['urls']
	name=session['name']
	email=session['email']
	filenames_errors=None
	msg_company=None
	msg_cmdb1=None
	msg_cmdb2=None
	msgCIs=None
	msgSites=None
	msg=None
	msg_details=None
	msg2=None
	msg3=None
	msg4=None
	msg5=None
	msg6=None
	msg7=None
	msg8=None
	msg9=None
	msg10=None
	msg11=None
	msg12=None
	msg13=None
	#msg14=None
	msg15=None
	#msg16=None
	msg17=None
	msg18=None
	add_msg_cis=''
	add_msg_sites=''
	content_mis_fields=''
	content_cmdb=''
	content_errors_Cis=''
	content_warnings_Cis=''
	content_summary_Cis=''
	content_errors_Sites=''
	content_warnings_Sites=''
	content_summary_Sites=''
	ID_FOLDER=session['filename']
	UPLOAD_FOLDER=ID_FOLDER + '/Files_to_validate/'
	DOWNLOAD_FOLDER=ID_FOLDER+'/Report/'
	filenames = []
	msg_company=ID_FOLDER.split('_')[0]
	instance='PROD'
	startTime = dt.datetime.now()
	process_data.process_file(path=UPLOAD_FOLDER,company=ID_FOLDER.split('_')[0],report=DOWNLOAD_FOLDER,instance=instance)
	filenames = [f for f in os.listdir(DOWNLOAD_FOLDER) if f.endswith(('csv'))]
	filenames_errors = [f for f in os.listdir(DOWNLOAD_FOLDER) if f.endswith(('.xlsx'))]
	if len(filenames_errors)>0:
		xl = pd.ExcelFile(DOWNLOAD_FOLDER+filenames_errors[0])
		if xl.sheet_names[0]=='Sheet1':
			filenames_errors=None
		else:
			None
	else:
		None
	#if len(filenames_errors)>0:
	#	report_issues = load_workbook(DOWNLOAD_FOLDER+filenames_errors[0])
	#	sheets_sites = report_issues.sheetnames
	#	sheets_issues=[x for x in sheets_sites if any(c.isdigit() for c in x)]
	#	for i in range(len(sheets_issues)):
	#		report_issues[sheets_issues[i]].sheet_properties.tabColor = 'FB0606'
	#	report_issues.save(DOWNLOAD_FOLDER+filenames_errors[0])
	#else:
	#	None
	#msg_time='Time spent in generating the validation report: '+str(dt.datetime.now() - startTime)
	msg_time=(dt.datetime.now() - startTime).seconds

	msg_company=ID_FOLDER.split('_')[0]
	if 'Mismatched_fields.txt' in os.listdir(DOWNLOAD_FOLDER):
		text_mis_fields=open(DOWNLOAD_FOLDER+'Mismatched_fields.txt', 'r+',encoding='utf8')
		content_mis_fields = text_mis_fields.read()
		text_mis_fields.close()
		msg17='WRONG TEMPLATE USED OR FIELDS ARE MISSING FROM ORIGINAL TEMPLATE.'
		msg18='PLEASE USE THE CORRECT TEMPLATES PROVIDED IN HOMEPAGE.'
	else:
		content_mis_fields=''
		if 'summary_CMDB.txt' in os.listdir(DOWNLOAD_FOLDER):
			text_cmdb=open(DOWNLOAD_FOLDER+'summary_CMDB.txt', 'r+',encoding='utf8')
			content_cmdb = text_cmdb.read()
			text_cmdb.close()
			msg_cmdb1=ID_FOLDER.split('_')[0].upper() + ' Current CMDB size in ITSM: '.upper()
		else:
			msg_cmdb2='Missing CMDB inventory'.upper()
		msgCIs="CI'S VALIDATION:"
		if 'errorsCIs.txt' in os.listdir(DOWNLOAD_FOLDER):
			text_errors_Cis=open(DOWNLOAD_FOLDER+'errorsCIs.txt', 'r+',encoding='utf8')
			content_errors_Cis = text_errors_Cis.read()
			text_errors_Cis.close()
			if len(content_errors_Cis)>1:
				msg='ERRORS'
				msg_details='Errors found in CIs data. Please download the report and check the sheets.'
			else:
				msg2='Your CIs data has no errors and is ready to be uploaded.'
		else:
			msg3='CIS FILE NOT FOUND.'
			add_msg_cis='(If you still want to validate CIs, make sure you enter the correct file or the correct sheetname file - cis).'

		if 'warningsCIs.txt' in os.listdir(DOWNLOAD_FOLDER):
			text_warnings_Cis=open(DOWNLOAD_FOLDER+'warningsCIs.txt', 'r+',encoding='utf8')
			content_warnings_Cis = text_warnings_Cis.read()
			text_warnings_Cis.close()
			if len(content_warnings_Cis)>1:
				msg7='WARNINGS:'
			else:
				content_warnings_Cis=''
				msg8='YOUR CIS DATA HAS NO WARNINGS.'
		else:
			None

		if 'summaryCIs.txt' in os.listdir(DOWNLOAD_FOLDER):
			text_summary_Cis=open(DOWNLOAD_FOLDER+'summaryCIs.txt', 'r+',encoding='utf8')
			content_summary_Cis = text_summary_Cis.read()
			text_summary_Cis.close()
			msg15='SUMMARY:'
		else:
			None
		msgSites="SITES VALIDATION"
		if 'errorsSites.txt' in os.listdir(DOWNLOAD_FOLDER):
			text_errors_Sites=open(DOWNLOAD_FOLDER+'errorsSites.txt', 'r+',encoding='utf8')
			content_errors_Sites = text_errors_Sites.read()
			text_errors_Sites.close()
			if len(content_errors_Sites)>1:
				msg4='Errors found in Sites data. Please download the report and check the sheets.'
			else:
				msg5='Your Sites data has no errors and is ready to be uploaded.'
				
		else:
			msg6='SITES FILE NOT FOUND'
			add_msg_sites='(If you still want to validate Sites, make sure you enter the correct file or the correct sheetname file - sites).'

		
		if 'warningsSites.txt' in os.listdir(DOWNLOAD_FOLDER):
			text_warnings_Sites=open(DOWNLOAD_FOLDER+'warningsSites.txt', 'r+',encoding='utf8')
			content_warnings_Sites = text_warnings_Sites.read()
			text_warnings_Sites.close()
			if len(content_warnings_Sites)>1:
				msg10='WARNINGS:'
			else:
				msg11='YOUR SITES DATA HAS NO WARNINGS.'
		else:
			None
			

		if 'summarySites.txt' in os.listdir(DOWNLOAD_FOLDER):
			text_summary_Sites=open(DOWNLOAD_FOLDER+'summarySites.txt', 'r+',encoding='utf8')
			content_summary_Sites = text_summary_Sites.read()
			text_summary_Sites.close()
			msg13='SUMMARY:'
		else:
			None
	return render_template('multi_files_upload.html',
		filenames=filenames,
		filenames_errors=filenames_errors,
		text_cmdb=content_cmdb,
		msg_time=msg_time,
		add_msg_cis=add_msg_cis,
		add_msg_sites=add_msg_sites,
		text_mis_fields=content_mis_fields,
		text_errors_Cis=content_errors_Cis,
		text_errors_Sites=content_errors_Sites,
		text_warnings_Cis=content_warnings_Cis,
		text_warnings_Sites=content_warnings_Sites,
		text_summary_Cis=content_summary_Cis,
		text_summary_Sites=content_summary_Sites,
		msg_company=msg_company,msg_cmdb1=msg_cmdb1,msg_cmdb2=msg_cmdb2,msgCIs=msgCIs,msgSites=msgSites,msg=msg,msg_details=msg_details,msg2=msg2,msg3=msg3,msg4=msg4,msg5=msg5,msg6=msg6,msg7=msg7,
		msg8=msg8,msg9=msg9,msg10=msg10,msg11=msg11,msg12=msg12,msg13=msg13,msg15=msg15,msg17=msg17,msg18=msg18)


@application.route('/report/<filename>')
@login_required
def uploaded_file(filename):
	ID_FOLDER=session['filename']
	DOWNLOAD_FOLDER=ID_FOLDER+'/Report/'
	return send_from_directory(DOWNLOAD_FOLDER,filename)




#####download report files
#@application.route('/return-file/')
#@login_required
#def return_file():
#	ID_FOLDER=session['filename']
#	DOWNLOAD_FOLDER=ID_FOLDER+'/Report/'
#	filename=os.listdir(DOWNLOAD_FOLDER)[0]
#	return send_file(DOWNLOAD_FOLDER+filename,attachment_filename=filename, as_attachment=True)
#
#@application.route('/templates')
#@login_required
#def file_downloads():
#	return render_template('multi_files_upload.html')

##################################################################
################convert to NOAM

@application.route('/noam_data', methods=['GET','POST'])
@login_required
def noam_data():
	msg_company=None
	msg3=None
	msg_to_many_files=None
	session['filename']=session['company']+'_'+str(uuid.uuid1())
	ID_FOLDER=session['filename']
	UPLOAD_FOLDER=ID_FOLDER + '/Files_to_validate/'
	msg_company=ID_FOLDER.split('_')[0]
	# Get the name of the uploaded files
	uploaded_files = request.files.getlist("file[]")
	for file in uploaded_files:
		# Check if the file is one of the allowed types/extensions
		if file and allowed_file(file.filename):
			# Make the filename safe, remove unsupported chars
			filename = secure_filename(file.filename)
			if not os.path.exists(ID_FOLDER):
				os.makedirs(ID_FOLDER)
				os.makedirs(UPLOAD_FOLDER)
			# Move the file form the temporal folder to the upload
			
			file.save(os.path.join(UPLOAD_FOLDER, filename))
			filenames=os.listdir(UPLOAD_FOLDER)
			if len(os.listdir(UPLOAD_FOLDER))>2:
				msg_to_many_files='THE NUMBER OF INPUT FILES SHOULD BE 1 (sites or cis) OR 2(sites and cis). PLEASE SELECT AGAIN THE INPUT FILES.'

			else:
				msg3='File(s) Successfully Submited. Click ont the button bellow to start the validation.'
		else:
			msg3='Please select a valid extension (.xls or .xlsx)'
	return render_template('index_NOAM_company.html',msg3=msg3,msg_company=msg_company,msg_to_many_files=msg_to_many_files)


@application.route('/noam_upload', methods=['GET'])
@login_required
def noam_upload():
	msg_company=None
	msg_miss_file=None
	ID_FOLDER=session['filename']
	msg_company=ID_FOLDER.split('_')[0]
	UPLOAD_FOLDER=ID_FOLDER + '/Files_to_validate/'
	DOWNLOAD_FOLDER=ID_FOLDER+'/Report/'
	os.makedirs(DOWNLOAD_FOLDER)
	
	# Get the name of the uploaded files
	if not os.path.exists(UPLOAD_FOLDER):
		msg_miss_file='PLEASE INSERT THE INPUT FILE(S) BEFORE RUN THE VALIDATION'
		return render_template('index_NOAM_company.html',msg_miss_file=msg_miss_file,msg_company=msg_company)
	else:
	#if len(os.listdir(UPLOAD_FOLDER))>0:
		process_noam_data.noam_files(file_path=UPLOAD_FOLDER,company=ID_FOLDER.split('_')[0],NOAM_report=DOWNLOAD_FOLDER)
		noam_filenames=os.listdir(DOWNLOAD_FOLDER)

		return render_template('noam_files_upload.html', noam_filenames=noam_filenames,msg_company=msg_company)


@application.route('/NOAM_Report/<filename>')
@login_required
def uploaded_NOAM_file(filename):
	ID_FOLDER=session['filename']
	DOWNLOAD_FOLDER=ID_FOLDER +'/Report/'
	return send_from_directory(DOWNLOAD_FOLDER,filename)

##########################################################
###RESOLUTION AND OPERATIONAL Categories valdidation
@application.route('/op_res_cats', methods=['GET','POST'])
@login_required
def op_res_cats_data():
	msg_company=None
	msg3=None
	msg_to_many_files=None
	session['filename']=session['company']+'_'+str(uuid.uuid1())
	ID_FOLDER=session['filename']
	msg_company=ID_FOLDER.split('_')[0]
	UPLOAD_FOLDER=ID_FOLDER + '/Files_to_validate/'
	uploaded_files = request.files.getlist("file[]")
	for file in uploaded_files:
		# Check if the file is one of the allowed types/extensions
		if file and allowed_file(file.filename):
			# Make the filename safe, remove unsupported chars
			filename = secure_filename(file.filename)
			
			if not os.path.exists(ID_FOLDER):
				os.makedirs(ID_FOLDER)
				os.makedirs(UPLOAD_FOLDER)
			# Move the file form the temporal folder to the upload
			
			file.save(os.path.join(UPLOAD_FOLDER, filename))
			filenames=os.listdir(UPLOAD_FOLDER)
			if len(os.listdir(UPLOAD_FOLDER))>2:
				msg_to_many_files='THE NUMBER OF INPUT FILES SHOULD BE 1 (OPS or RES CATS) OR 2(OPS and RES CATS). PLEASE SELECT AGAIN THE INPUT FILES.'

			else:
				msg3='File(s) Successfully Submited. Click ont the button bellow to start the validation.'
		else:
			msg3='Please select a valid extension (.xls or .xlsx)'
	return render_template('index_res_cats.html',msg3=msg3,msg_company=msg_company,msg_to_many_files=msg_to_many_files)



@application.route('/op_res_cats_upload', methods=['GET'])
@login_required
def op_res_cats_data_upload():
	user=session['username']
	msg_company=None
	msg=None
	msg2=None
	msg3=None
	add_msg_res=None
	msg_res=None
	msg4=None
	msg5=None
	msg6=None
	add_msg_op=None
	msg_ops=None
	msg_fields_res=None
	text_mis_fields_res=None
	msg_fields_ops=None
	text_mis_fields_ops=None
	#msg_fields_2=None
	ID_FOLDER=session['filename']
	msg_company=ID_FOLDER.split('_')[0]

	UPLOAD_FOLDER=ID_FOLDER + '/Files_to_validate/'
	DOWNLOAD_FOLDER=ID_FOLDER+'/Report/'
	os.makedirs(DOWNLOAD_FOLDER)
	#OP_RES_FOLDER=session['filename']
	#OP_RES_UPLOAD=OP_RES_FOLDER+'/op_res_cats_files/'
	#OP_RES_REPORT=OP_RES_FOLDER +'/Report/'
	# Get the name of the uploaded files
	if not os.path.exists(UPLOAD_FOLDER):
		msg_miss_file='PLEASE INSERT THE INPUT FILE(S) BEFORE RUN THE VALIDATION'
		return render_template('index_res_cats.html',msg_miss_file=msg_miss_file,msg_company=msg_company)
	else:
		###type of user form selection
		if user!='numartin' and user!='paulof':
			process_res_cats_test.op_res_cats_files(path=UPLOAD_FOLDER,company=ID_FOLDER.split('_')[0],op_res_cats_report=DOWNLOAD_FOLDER)
		else:
			process_res_cats.op_res_cats_files(path=UPLOAD_FOLDER,company=ID_FOLDER.split('_')[0],op_res_cats_report=DOWNLOAD_FOLDER)
		####filenames to dowonload	
		op_res_filenames=[f for f in os.listdir(DOWNLOAD_FOLDER) if f.endswith(('.xlsm','xlsx'))]
		content_res=''
		content_ops=''
		content_mis_fields_res=''
		content_mis_fields_ops=''
		###RES
		msg_res="RESOLUTION CATEGORIES VALIDATION:"
		if 'Mismatched_fields_res.txt' in os.listdir(DOWNLOAD_FOLDER):
			text_mis_fields_res=open(DOWNLOAD_FOLDER+'Mismatched_fields_res.txt', 'r+',encoding='utf8')
			content_mis_field_res = text_mis_fields_res.read()
			text_mis_fields_res.close()
			msg_fields_res='WRONG OR MISSING FIELDS FROM ORIGINAL RESCATS TEMPLATE.'
			#msg_fields_2='PLEASE USE THE CORRECT TEMPLATES PROVIDED IN HOMEPAGE.'
		else:
			content_mis_fields_res=''
			
			if 'res_issues.txt' in os.listdir(DOWNLOAD_FOLDER):
				print('Yes')
				text_res=open(DOWNLOAD_FOLDER+'res_issues.txt', 'r+',encoding='utf8')
				content_res = text_res.read()
				text_res.close()
				if len(content_res)>1:
					msg='Errors found in Resolution Categories. Please check the Catalogue.'
				else:
					msg2='Your Resolution Categories are correct and ready to be uploaded.'
			else:
				msg3='Resolution Categories not found.'
				add_msg_res='(If you still want to validate Resolution Categories, make sure you enter the correct file or the correct sheetname file - ResCat).'
		#######OPS
		msg_ops="OPERATIONAL CATEGORIES VALIDATION:"
		if 'Mismatched_fields_ops.txt' in os.listdir(DOWNLOAD_FOLDER):
			text_mis_fields_ops=open(DOWNLOAD_FOLDER+'Mismatched_fields_ops.txt', 'r+',encoding='utf8')
			content_mis_field_ops = text_mis_fields_ops.read()
			text_mis_fields_ops.close()
			msg_fields_ops='WRONG OR MISSING FIELDS FROM ORIGINAL OPCATS TEMPLATE.'
			#msg_fields_2='PLEASE USE THE CORRECT TEMPLATES PROVIDED IN HOMEPAGE.'
		else:
			content_mis_fields_ops=''	
			
			if 'op_issues.txt' in os.listdir(DOWNLOAD_FOLDER):
				text_ops=open(DOWNLOAD_FOLDER+'op_issues.txt', 'r+',encoding='utf8')
				content_ops = text_ops.read()
				text_ops.close()
				if len(content_ops)>1:
					msg4='Errors found in Operational Categories. Please check the Catalogue.'
				else:
					msg5='Your Operational Categories are correct and ready to be uploaded.'
			else:
				msg6='Operational Categories not found.'
				add_msg_op='(If you still want to validate Operational Categories, make sure you enter the correct file or the correct sheetname file - OpCat).'
		return render_template('res_cats_upload.html',
			msg_company=msg_company,
			msg_fields_res=msg_fields_res,
			text_mis_fields_res=content_mis_fields_res,
			msg_fields_ops=msg_fields_ops,
			text_mis_fields_ops=content_mis_fields_ops,
			op_res_filenames=op_res_filenames,
			text_res=content_res,
			text_ops=content_ops,
			msg_res=msg_res,
			msg_ops=msg_ops,
			msg=msg,msg2=msg2,msg3=msg3,add_msg_res=add_msg_res,msg4=msg4,msg5=msg5,msg6=msg6,add_msg_op=add_msg_op)


@application.route('/OP_RES_UPLOAD_Report/<filename>')
@login_required
def uploaded_RES_CATS_file(filename):
	ID_FOLDER=session['filename']
	DOWNLOAD_FOLDER=ID_FOLDER +'/Report/'
	return send_from_directory(DOWNLOAD_FOLDER,filename)

#############################################################
######ZTE split files
@application.route('/eia', methods=['GET','POST'])
@login_required
def eia_data():
	msg=None
	session['filename']=str(uuid.uuid1())
	EIA_FOLDER=session['filename']
	EIA_UPLOAD=EIA_FOLDER+'/eia_files/'
	EIA_REPORT=EIA_FOLDER +'/Report/'
	
	# Get the name of the uploaded files
	uploaded_files = request.files.getlist("file[]")
	for file in uploaded_files:
		# Check if the file is one of the allowed types/extensions
		if file and allowed_file(file.filename):
			# Make the filename safe, remove unsupported chars
			filename = secure_filename(file.filename)
			if not os.path.exists(EIA_FOLDER):
				os.makedirs(EIA_FOLDER)
				os.makedirs(EIA_UPLOAD)
				os.makedirs(EIA_REPORT)
			# Move the file form the temporal folder to the upload
			
			file.save(os.path.join(EIA_UPLOAD, filename))
			filenames=os.listdir(EIA_UPLOAD)
			msg=filenames
		else:
			msg='Please select a valid extension (.xls, .xlsx or .csv)'
	return render_template('index_eia.html',msg=msg)

@application.route('/eia_upload', methods=['GET'])
@login_required
def eia_upload():
	EIA_FOLDER=session['filename']
	EIA_UPLOAD=EIA_FOLDER+'/eia_files/'
	EIA_REPORT=EIA_FOLDER +'/Report/'
	# Get the name of the uploaded files
	if len(os.listdir(EIA_UPLOAD))>0:
		process_zte.split_zte_file(path=EIA_UPLOAD,report=EIA_REPORT)
		eia_filenames=os.listdir(EIA_REPORT)
	return render_template('eia_upload.html', eia_filenames=eia_filenames)



@application.route('/EIA_Report/<filename>')
@login_required
def uploaded_EIA_file(filename):
	EIA_FOLDER=session['filename']
	EIA_REPORT=EIA_FOLDER +'/Report/'
	return send_from_directory(EIA_REPORT,filename)

####CHANGE MODULE

@application.route('/chg_mod', methods=['GET'])
@login_required
def chg_mod():
	msg='TO BE IMPLEMENTED'
	return render_template('chg_mod.html',msg=msg)

###RAP
@application.route('/rap', methods=['GET','POST'])
@login_required
def rap():
	session['filename']=session['company']+'_'+str(uuid.uuid1())
	ID_FOLDER=session['filename']
	msg_company=ID_FOLDER.split('_')[0]
	UPLOAD_FOLDER=ID_FOLDER + '/Files_to_validate/'

	msg=None
	if request.method == 'POST':
		if 'file' not in request.files:
			print('No file attached in request')
			return redirect(request.url)
		file = request.files['file']
		if file.filename == '':
			print('No file selected')
			return redirect(request.url)
		if file and allowed_file(file.filename):
			filename = secure_filename(file.filename)
			if not os.path.exists(ID_FOLDER):
				os.makedirs(ID_FOLDER)
				os.makedirs(UPLOAD_FOLDER)
			file.save(os.path.join(UPLOAD_FOLDER, filename))
			msg=filename
		else:
			msg='Please select a valid extension (.xls or .xlsx)'
	return render_template('rap.html',msg=msg)


@application.route('/rap_upload', methods=['POST'])
@login_required
def rap_upload():
	cis_itsm=pd.DataFrame()
	msg=None
	ID_FOLDER=session['filename']
	UPLOAD_FOLDER=ID_FOLDER + '/Files_to_validate/'
	DOWNLOAD_FOLDER=ID_FOLDER+'/Report/'
	os.makedirs(DOWNLOAD_FOLDER)
	# Get the name of the uploaded files
	if len(os.listdir(UPLOAD_FOLDER))>0:

		process_rap.process_rap_file(path=UPLOAD_FOLDER,company=ID_FOLDER.split('_')[0],report=DOWNLOAD_FOLDER)
		process_cmdb_inventory.call_cmdb_inventory(company=ID_FOLDER.split('_')[0],report=DOWNLOAD_FOLDER)

		rap_filenames=os.listdir(DOWNLOAD_FOLDER)
		print(rap_filenames)
		
		if 'issues.txt' in os.listdir(DOWNLOAD_FOLDER):
			text_rap=open(DOWNLOAD_FOLDER+'issues.txt', 'r+',encoding='utf8')
			content_rap = text_rap.read()
			text_rap.close()
			msg='ERRORS:'
		else:
			None
	return render_template('rap_upload.html', 
		rap_filenames=rap_filenames,
		text_rap=content_rap,msg=msg)

@application.route('/RAP_Report/<filename>')
@login_required
def uploaded_RAP_file(filename):
	ID_FOLDER=session['filename']
	DOWNLOAD_FOLDER=ID_FOLDER +'/Report/'
	return send_from_directory(DOWNLOAD_FOLDER,filename)


###########################convert rap to noam

##################################################################
################convert to NOAM

@application.route('/noam_rap', methods=['GET','POST'])
@login_required
def noam_rap():
	msg_company=None
	msg3=None
	session['filename']=session['company']+'_'+str(uuid.uuid1())
	ID_FOLDER=session['filename']
	UPLOAD_FOLDER=ID_FOLDER + '/Files_to_validate/'
	msg_company=ID_FOLDER.split('_')[0]
	# Get the name of the uploaded files
	uploaded_files = request.files.getlist("file[]")
	for file in uploaded_files:
		# Check if the file is one of the allowed types/extensions
		if file and allowed_file(file.filename):
			# Make the filename safe, remove unsupported chars
			filename = secure_filename(file.filename)
			if not os.path.exists(ID_FOLDER):
				os.makedirs(ID_FOLDER)
				os.makedirs(UPLOAD_FOLDER)
			# Move the file form the temporal folder to the upload
			
			file.save(os.path.join(UPLOAD_FOLDER, filename))
			filenames=os.listdir(UPLOAD_FOLDER)
			msg3=filenames
		else:
			msg3='Please select a valid extension (.xls or .xlsx)'
	return render_template('index_NOAM_rap.html',msg3=msg3,msg_company=msg_company)


@application.route('/rap_noam_upload', methods=['GET'])
@login_required
def noam_rap_upload():
	msg_company=None
	msg_miss_file=None
	ID_FOLDER=session['filename']
	msg_company=ID_FOLDER.split('_')[0]
	UPLOAD_FOLDER=ID_FOLDER + '/Files_to_validate/'
	DOWNLOAD_FOLDER=ID_FOLDER+'/Report/'
	os.makedirs(DOWNLOAD_FOLDER)
	
	# Get the name of the uploaded files
	if not os.path.exists(UPLOAD_FOLDER):
		msg_miss_file='PLEASE INSERT THE INPUT FILE(S) BEFORE RUN THE VALIDATION'
		return render_template('index_NOAM_rap.html',msg_miss_file=msg_miss_file,msg_company=msg_company)
	else:
	#if len(os.listdir(UPLOAD_FOLDER))>0:
		process_rap_noam.noam_rap(file_path=UPLOAD_FOLDER,company=ID_FOLDER.split('_')[0],NOAM_report=DOWNLOAD_FOLDER)
		noam_filenames=os.listdir(DOWNLOAD_FOLDER)

		return render_template('noam_rap_upload.html', noam_filenames=noam_filenames,msg_company=msg_company)


@application.route('/NOAM_rap_Report/<filename>')
@login_required
def uploaded_NOAM_rap(filename):
	ID_FOLDER=session['filename']
	DOWNLOAD_FOLDER=ID_FOLDER +'/Report/'
	return send_from_directory(DOWNLOAD_FOLDER,filename)





###update CMDB####

####update prodCat############
@application.route('/update_cmdb', methods=['GET','POST'])
@login_required
def update_cmdb():
	msg3=None
	session['filename']=session['company']+'_'+str(uuid.uuid1())
	ID_FOLDER=session['filename']
	TEMP_FOLDER=ID_FOLDER +'/update_cmdb/'
	msg_company=ID_FOLDER.split('_')[0]
	uploaded_files = request.files.getlist("file[]")
	for file in uploaded_files:
		# Check if the file is one of the allowed types/extensions
		if file and allowed_file(file.filename):
			
			# Make the filename safe, remove unsupported chars
			filename = secure_filename(file.filename)
			if not os.path.exists(TEMP_FOLDER):
				os.makedirs(ID_FOLDER)
				os.makedirs(TEMP_FOLDER)
			# Move the file form the temporary folder to the upload
			file.save(os.path.join(TEMP_FOLDER, filename))
			filenames=os.listdir(TEMP_FOLDER)
			msg3=filenames
		else:
			msg3='Please select a valid extension (.xls(x) or .csv)'

	return render_template('update_cmdb.html',msg3=msg3,msg_company=msg_company)

#######upload prod_cat

@application.route('/cmdb_upload', methods=['GET'])
@login_required
def cmdb_upload():
	ID_FOLDER=session['filename']
	TEMP_FOLDER=ID_FOLDER +'/update_cmdb/'
	msg_company=ID_FOLDER.split('_')[0]
	process_cmdb_update.cmdb_update(path=TEMP_FOLDER,company=ID_FOLDER.split('_')[0])
	#cmdb_filenames=[f for f in os.listdir(DOWNLOAD_FOLDER) if f.endswith(('.xlsx','csv'))]
	return render_template('upload_cmdb.html',msg_company=msg_company)
#
#
#
#@application.route('/report_prod/<filename>')
#@login_required
#def uploaded_prod(filename):
#	DOWNLOAD_FOLDER= 'Prod_Cats_V2/'
#	return send_from_directory(DOWNLOAD_FOLDER,filename)
######################################################################





@application.route('/logout', methods=['GET'])
@login_required
def logout():
	#ID_FOLDER=session['filename']
	#if os.path.exists(ID_FOLDER):
	#	shutil.rmtree(ID_FOLDER)
	logout_user()
	return redirect('login')


# create login manager
login_manager = LoginManager()
# init login manager on application
login_manager.init_app(application)
# set login view
login_manager.login_view = 'login_get'
# callback to reload the user object
@login_manager.user_loader
def load_user(userid):
	return USERS[userid]

if __name__=='__main__':
	application.run(debug=True,threaded = True)