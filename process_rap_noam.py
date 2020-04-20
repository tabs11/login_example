import pandas as pd
import numpy as np
import os
import re
from openpyxl import load_workbook
import itertools
import datetime as dt
import glob
def noam_rap(file_path,company,NOAM_report):
	noam_rap = load_workbook('CMDB_templates/rap.xlsm',read_only=False, keep_vba=True)
	sheets_rap = noam_rap.sheetnames
	w_sheet1_rap=noam_rap[sheets_rap[1]]
	##files to upload
	column_names=['RA Bulk Upload Template']
	rap=pd.read_excel(glob.glob(file_path+'*')[0],list(set(pd.ExcelFile(glob.glob(file_path+'*')[0]).sheet_names).intersection(column_names))[0])
	rap.fillna('',inplace=True)
	rap['Requested Start Date (DD/MM/YYYY)']=pd.to_datetime(rap['Requested Start Date (DD/MM/YYYY)']).apply(lambda x: x.strftime('%m/%d/%Y'))
	for i in range(np.shape(rap)[0]):
		w_sheet1_rap['A' +str(4+i)]=rap['SA Number'][i]
		w_sheet1_rap['B' +str(4+i)]=rap['Company'][i]
		w_sheet1_rap['C' +str(4+i)]=rap['Short Description'][i]
		w_sheet1_rap['D' +str(4+i)]=rap['Description'][i]
		w_sheet1_rap['E' +str(4+i)]='Active'
		w_sheet1_rap['F' +str(4+i)]=rap['Frequency'][i]
		w_sheet1_rap['G' +str(4+i)]=rap['Recommended Frequency'][i]
		w_sheet1_rap['H' +str(4+i)]=rap['RA Trigger'][i]
		w_sheet1_rap['I' +str(4+i)]=rap['Window'][i]
		w_sheet1_rap['J' +str(4+i)]=rap['Intervention Type'][i]
		w_sheet1_rap['B' +str(4+i)]=rap['Service Impact'][i]
		w_sheet1_rap['S' +str(4+i)]=rap['Requested Start Date (DD/MM/YYYY)'][i]
		w_sheet1_rap['T' +str(4+i)]=rap['Requested End Date (MM/DD/YYYY)'][i]
		if rap['CI Name'].isnull().all():
			print('NO CIs to check')
		else:
			w_sheet1_rap['AC' +str(4+i)]=rap['CI Name'][i]
		if rap['Site Name'].isnull().all():
			print('NO Sites to check')
		else:
			w_sheet1_rap['AB' +str(4+i)]=rap['Site Name'][i]
			
	noam_rap.save(filename = NOAM_report + company + '_RAP_Noam_' + dt.datetime.now().strftime("%Y-%m-%d %H-%M-%S") +'.xlsm')
	

