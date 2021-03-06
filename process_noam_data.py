import pandas as pd
import numpy as np
import os
import re
from openpyxl import load_workbook
import itertools
import datetime as dt
import glob
def noam_files(file_path,company,NOAM_report):
	noam_sites = load_workbook('CMDB_templates/Location_original.xlsm',read_only=False, keep_vba=True)
	sheets_sites = noam_sites.sheetnames
	w_sheet1_sites=noam_sites[sheets_sites[1]]
	w_sheet2_sites=noam_sites[sheets_sites[2]]
	w_sheet3_sites=noam_sites[sheets_sites[3]]
	w_sheet4_sites=noam_sites[sheets_sites[4]]
	w_sheet5_sites=noam_sites[sheets_sites[5]]
	##noam cis
	noam_cis = load_workbook('CMDB_templates/Transactional_CI_original.xlsm',read_only=False, keep_vba=True)
	sheets_cis = noam_cis.sheetnames
	w_sheet2_cis=noam_cis[sheets_cis[2]]
	##files to upload
	sheets=[]
	cis=[]
	sites=[]
	#for j in range(len(glob.glob(file_path+'/*'))):
	#	for k in range(len(pd.ExcelFile(glob.glob(file_path+'*')[j]).sheet_names)):
	#		sheets.append(pd.read_excel(glob.glob(file_path+'*')[j],pd.ExcelFile(glob.glob(file_path+'*')[j]).sheet_names[k]))     
	#
	column_names=['sites','cis','Site Data Template','CI Data Template Comp Syst']
	for j in range(len(glob.glob(file_path+'/*'))):
		if glob.glob(file_path+'/*')[j].endswith(('.xls','.xlsx')):
			for k in range(len(list(set(pd.ExcelFile(glob.glob(file_path+'*')[j]).sheet_names).intersection(column_names)))):
				sheets.append(pd.read_excel(glob.glob(file_path+'*')[j],list(set(pd.ExcelFile(glob.glob(file_path+'*')[j]).sheet_names).intersection(column_names))[k]))
			#files=pd.read_excel(glob.glob(path+'*')[j],sheet_name=None)
			#for frame in files.keys():
			#	sheets.append(files[frame])
		elif glob.glob(file_path+'/*')[j].endswith('.csv'):
			sheets.append(pd.read_csv(glob.glob(file_path+'*')[j],sep=";",encoding='ISO-8859-1'))
		else:
			None


	for j in range(len(sheets)):
		if (~sheets[j].columns.str.contains('CI N',case=False).any()) & (sheets[j].columns.str.contains('SITE N|SITE+|SITE*',case=False).any()):
			sites.append(sheets[j])
		else:
			None
		#CIS
		if sheets[j].columns.str.contains('CI N',case=False).any():
			cis.append(sheets[j])
		else:
			None

	if len(sites)>0:
		sites=sites[0]
		sites.fillna('',inplace=True)
		#sites_region=sites[['Region','Company','Status']]
		#sites_groups=sites[['Site Group','Description','Region','Company','Status']]
		#sites_region.loc[sites_region['Region'].duplicated()] = ''
		#sites_region.sort_values([sites_region.columns[0]],ascending=False).reset_index(drop=True,inplace=True)
		#sites=sites.reindex(sites_region.index)

		#site_alias=sites.copy()
		#sites.drop('Site Alias',axis=1, inplace=True)
		#sites.drop_duplicates(inplace=True)
		for i in range(np.shape(sites)[0]):
		#site
			w_sheet1_sites['A' +str(4+i)]=sites['Site Name'][i]
			w_sheet4_sites['A' +str(4+i)]=sites['Site Name'][i]
			w_sheet5_sites['A' +str(4+i)]=sites['Site Name'][i]
			w_sheet5_sites['B' +str(4+i)]=sites['Site Alias'][i]
			##street
		
			w_sheet1_sites['B' +str(4+i)]=sites['Street'][i]
			
			##country
			w_sheet1_sites['C' +str(4+i)]=sites['Country'][i]
			##city
			w_sheet1_sites['E' +str(4+i)]=sites['City'][i]
			##Location ID
			w_sheet1_sites['O' +str(4+i)]=sites['Location ID'][i]
			##Description
			w_sheet1_sites['P' +str(4+i)]=sites['Description'][i]
		
			w_sheet3_sites['D' +str(4+i)]=sites['Description'][i]
		
			##Additional Sites Details
			w_sheet1_sites['Q' +str(4+i)]=sites['Additional Site Details'][i]
		
			
			##status
			if (sites['Status']=='').any():
				w_sheet1_sites['R' +str(4+i)]='Enabled'
				w_sheet2_sites['C' +str(4+i)]='Enabled'
				w_sheet3_sites['E' +str(4+i)]='Enabled'
				w_sheet4_sites['E' +str(4+i)]='Enabled'
			else:
				w_sheet1_sites['R' +str(4+i)]=sites['Status'][i]
				w_sheet2_sites['C' +str(4+i)]=sites['Status'][i]
				w_sheet3_sites['E' +str(4+i)]=sites['Status'][i]
				w_sheet4_sites['E' +str(4+i)]=sites['Status'][i]
		#
# 		
			#Latitude
			w_sheet1_sites['S' +str(4+i)]=sites['Latitude'][i]
			#LON
			w_sheet1_sites['T' +str(4+i)]=sites['Longitude'][i]
			#Maintenance cirlce name
			w_sheet1_sites['Y' +str(4+i)]=sites['Maintenance Circle Name'][i]
			#SITE TYPE
			w_sheet1_sites['Z' +str(4+i)]=sites['Site Type'][i]
			##reg
			w_sheet2_sites['A' +str(4+i)]=sites['Region'][i]
			w_sheet3_sites['C' +str(4+i)]=sites['Region'][i]
			
			w_sheet4_sites['C' +str(4+i)]=sites['Region'][i]
			#company
			w_sheet2_sites['B' +str(4+i)]=sites['Company'][i]
			w_sheet3_sites['B' +str(4+i)]=sites['Company'][i]
			w_sheet4_sites['B' +str(4+i)]=sites['Company'][i]
			#Site Group
			w_sheet3_sites['A' +str(4+i)]=sites['Site Group'][i]
			w_sheet4_sites['D' +str(4+i)]=sites['Site Group'][i]
	
		noam_sites.save(filename = NOAM_report + company + '_sites_Noam_' + dt.datetime.now().strftime("%Y-%m-%d %H-%M-%S") +'.xlsm')
	else:
		None
	if len(cis)>0:
		cis=cis[0]
		cis.fillna('',inplace=True)
		for i in range(np.shape(cis)[0]):
			#print(i)
			#static fields
			w_sheet2_cis['DJ' +str(4+i)]='migrator'
			w_sheet2_cis['DN' +str(4+i)]='Computer System'
			w_sheet2_cis['DT' +str(4+i)]='Computer System'
			w_sheet2_cis['DU' +str(4+i)]='BMC_COMPUTERSYSTEM'

			#CI type
			w_sheet2_cis['CL' +str(4+i)]='BMC_COMPUTERSYSTEM'
			##Product type
			w_sheet2_cis['CD' +str(4+i)]='Hardware'
			####Supported
			w_sheet2_cis['AC' +str(4+i)]='Yes'
			###CI Name
			w_sheet2_cis['D' +str(4+i)]=cis['CI Name'][i]

			##CI ID
			#w_sheet2_cis.write(4+i,11,cis.filter(regex=re.compile('CI ID',re.IGNORECASE)).iloc[:,0][i])
			w_sheet2_cis['L' +str(4+i)]=cis['CI ID'][i]
			###CI Description
			w_sheet2_cis['G' +str(4+i)]=cis['CI Description'][i]

			####Status
			if (cis['Status']=='').any():
				w_sheet2_cis['AP' +str(4+i)]='Deployed'
			else:
				w_sheet2_cis['AP' +str(4+i)]=cis['Status'][i]		
			####Tier 1
			w_sheet2_cis['B' +str(4+i)]=cis['Tier 1'][i]
			####Tier 2
			w_sheet2_cis['H' +str(4+i)]=cis['Tier 2'][i]
			####Tier 3
			w_sheet2_cis['J' +str(4+i)]=cis['Tier 3'][i]
			###Product Name
			#w_sheet2_cis.write(4+i,13,cis['Product Name+'][i])
			w_sheet2_cis['N' +str(4+i)]=cis['Product Name'][i]

			####Manufacturer
			w_sheet2_cis['V' +str(4+i)]=cis['Manufacturer'][i]

			####System Role
			w_sheet2_cis['AN' +str(4+i)]=cis['System Role'][i]

			###Priority
			if (cis['Priority']=='').any():
				w_sheet2_cis['BB' +str(4+i)]='PRIORITY_5'
			else:
				w_sheet2_cis['BB' +str(4+i)]=cis['Priority'][i].upper()

			###Additional Information
			w_sheet2_cis['BE' +str(4+i)]=cis['Additional Information'][i]
			##
			###Region
			w_sheet2_cis['AL' +str(4+i)]=cis['Region'][i]
			##
			###Site Group
			w_sheet2_cis['AR' +str(4+i)]=cis['Site Group'][i]
			##
			###Site
			w_sheet2_cis['AT' +str(4+i)]=cis['Site'][i]
			##
			###Tag Number
			w_sheet2_cis['S' +str(4+i)]=cis['Tag Number'][i]

			##
			###Model Version
			###w_sheet2_cis.write(4+i,19,cis[0].filter(regex=re.compile('Version',re.IGNORECASE)).iloc[:,0][i])
			##
			###DNS
			w_sheet2_cis['AJ' +str(4+i)]=cis['DNS Host Name'][i]

			##
			###Domain
			w_sheet2_cis['AQ' +str(4+i)]=cis['Domain'][i]


			##
			###num cells
			#if cis.columns.str.contains('cell',case=False).any():
			#	w_sheet2_cis['DW' +str(4+i)]=cis['NbrCells'][i]
			#else:
			#	print('No cells')
		noam_cis.save(filename = NOAM_report + company +'_CIs_Noam_'+ dt.datetime.now().strftime("%Y-%m-%d %H-%M-%S") + '.xlsm')
		#wb_cis.save(NOAM_report + company +'_cis_Noam.xls')
	else:
		None
#def noam_files(file_path,company,NOAM_report):
#
#	noam_sites = load_workbook('CMDB_templates/Location_original.xlsm',read_only=False, keep_vba=True)
#	sheets_sites = noam_sites.sheetnames
#	w_sheet1_sites=noam_sites[sheets_sites[1]]
#	w_sheet2_sites=noam_sites[sheets_sites[2]]
#	w_sheet3_sites=noam_sites[sheets_sites[3]]
#	w_sheet4_sites=noam_sites[sheets_sites[4]]
#	w_sheet5_sites=noam_sites[sheets_sites[5]]
#	##noam cis
#	noam_cis = load_workbook('CMDB_templates/Transactional_CI_original.xlsm',read_only=False, keep_vba=True)
#	sheets_cis = noam_cis.sheetnames
#	w_sheet2_cis=noam_cis[sheets_cis[2]]
#	##files to upload
#	sheets=[]
#	cis=[]
#	sites=[]
#	for j in range(len(glob.glob(file_path+'/*'))):
#		for k in range(len(pd.ExcelFile(glob.glob(file_path+'*')[j]).sheet_names)):
#			sheets.append(pd.read_excel(glob.glob(file_path+'*')[j],pd.ExcelFile(glob.glob(file_path+'*')[j]).sheet_names[k]))     
#	
#	for j in range(len(sheets)):
#		if (~sheets[j].columns.str.contains('CI N',case=False).any()) & (sheets[j].columns.str.contains('SITE N|SITE+|SITE*',case=False).any()):
#			sites.append(sheets[j])
#		else:
#			None
#		#CIS
#		if sheets[j].columns.str.contains('CI N',case=False).any():
#			cis.append(sheets[j])
#		else:
#			None
#
#
#	if len(sites)>0:
#		sites[0].fillna('',inplace=True)
#		for i in range(np.shape(sites[0])[0]):
#			#site
#			w_sheet1_sites['A' +str(4+i)]=sites[0].filter(regex=re.compile('SITE N',re.IGNORECASE)).iloc[:,0].values[i]
#			w_sheet4_sites['A' +str(4+i)]=sites[0].filter(regex=re.compile('SITE N',re.IGNORECASE)).iloc[:,0].values[i]
#			w_sheet5_sites['A' +str(4+i)]=sites[0].filter(regex=re.compile('SITE N',re.IGNORECASE)).iloc[:,0].values[i]
#			w_sheet5_sites['B' +str(4+i)]=sites[0].filter(regex=re.compile('ALIAS',re.IGNORECASE)).iloc[:,0].values[i]
#			##street
#
#			w_sheet1_sites['B' +str(4+i)]=sites[0].filter(regex=re.compile('STREET',re.IGNORECASE)).iloc[:,0].values[i]
#			
#			##country
#			w_sheet1_sites['C' +str(4+i)]=sites[0].filter(regex=re.compile('COUNTRY',re.IGNORECASE)).iloc[:,0].values[i]
#			##city
#			w_sheet1_sites['E' +str(4+i)]=sites[0].filter(regex=re.compile('CITY',re.IGNORECASE)).iloc[:,0].values[i]
#			##Location ID
#			w_sheet1_sites['O' +str(4+i)]=sites[0].filter(regex=re.compile('Location',re.IGNORECASE)).iloc[:,0].values[i]
#			##Description
#			w_sheet1_sites['P' +str(4+i)]=sites[0].filter(regex=re.compile('Descr',re.IGNORECASE)).iloc[:,0].values[i]
#
#			w_sheet3_sites['D' +str(4+i)]=sites[0].filter(regex=re.compile('Descr',re.IGNORECASE)).iloc[:,0].values[i]
#
#			##Additional Sites Details
#			w_sheet1_sites['Q' +str(4+i)]=sites[0].filter(regex=re.compile('Additional',re.IGNORECASE)).iloc[:,0].values[i]
#
#			
#			#status
#			if (sites[0].filter(regex=re.compile('STATUS',re.IGNORECASE)).iloc[:,0]=='').any():
#				w_sheet1_sites['R' +str(4+i)]='Enabled'
#				w_sheet2_sites['C' +str(4+i)]='Enabled'
#				
#				w_sheet3_sites['E' +str(4+i)]='Enabled'
#				w_sheet4_sites['E' +str(4+i)]='Enabled'
#			else:
#				w_sheet1_sites['R' +str(4+i)]=sites[0].filter(regex=re.compile('STATUS',re.IGNORECASE)).iloc[:,0].values[i]
#				w_sheet2_sites['C' +str(4+i)]=sites[0].filter(regex=re.compile('STATUS',re.IGNORECASE)).iloc[:,0].values[i]
#
#				w_sheet3_sites['E' +str(4+i)]=sites[0].filter(regex=re.compile('STATUS',re.IGNORECASE)).iloc[:,0].values[i]
#
#				w_sheet4_sites['E' +str(4+i)]=sites[0].filter(regex=re.compile('STATUS',re.IGNORECASE)).iloc[:,0].values[i]
#
#			
#			#Latitude
#			w_sheet1_sites['S' +str(4+i)]=sites[0].filter(regex=re.compile('LAT',re.IGNORECASE)).iloc[:,0].values[i]
#
#			#LON
#			w_sheet1_sites['T' +str(4+i)]=sites[0].filter(regex=re.compile('LON',re.IGNORECASE)).iloc[:,0].values[i]
#			#Maintenance cirlce name
#			w_sheet1_sites['Y' +str(4+i)]=sites[0].filter(regex=re.compile('CIRCLE',re.IGNORECASE)).iloc[:,0].values[i]
#			#SITE TYPE
#			w_sheet1_sites['Z' +str(4+i)]=sites[0].filter(regex=re.compile('TYPE',re.IGNORECASE)).iloc[:,0].values[i]
#			##reg
#			w_sheet2_sites['A' +str(4+i)]=sites[0].filter(regex=re.compile('REGION',re.IGNORECASE)).iloc[:,0].values[i]
#			w_sheet3_sites['C' +str(4+i)]=sites[0].filter(regex=re.compile('REGION',re.IGNORECASE)).iloc[:,0].values[i]
#
#			w_sheet4_sites['C' +str(4+i)]=sites[0].filter(regex=re.compile('REGION',re.IGNORECASE)).iloc[:,0].values[i]
#			#company
#			w_sheet2_sites['B' +str(4+i)]=sites[0].filter(regex=re.compile('COMPANY',re.IGNORECASE)).iloc[:,0].values[i]
#			w_sheet3_sites['B' +str(4+i)]=sites[0].filter(regex=re.compile('COMPANY',re.IGNORECASE)).iloc[:,0].values[i]
#			w_sheet4_sites['B' +str(4+i)]=sites[0].filter(regex=re.compile('COMPANY',re.IGNORECASE)).iloc[:,0].values[i]
#			#Site Group
#			w_sheet3_sites['A' +str(4+i)]=sites[0].filter(regex=re.compile('GROUP',re.IGNORECASE)).iloc[:,0].values[i]
#			w_sheet4_sites['D' +str(4+i)]=sites[0].filter(regex=re.compile('GROUP',re.IGNORECASE)).iloc[:,0].values[i]
#	
#		noam_sites.save(filename = NOAM_report + company + '_sites_Noam_' + dt.datetime.now().strftime("%Y-%m-%d %H-%M-%S") +'.xlsm')
#	else:
#		None
#	if len(cis)>0:
#		cis=cis[0]
#		cis.fillna('',inplace=True)
#		for i in range(np.shape(cis)[0]):
#			#static fields
#			w_sheet2_cis['DJ' +str(4+i)]='migrator'
#			w_sheet2_cis['DN' +str(4+i)]='Computer System'
#			w_sheet2_cis['DT' +str(4+i)]='Computer System'
#			w_sheet2_cis['DU' +str(4+i)]='BMC_COMPUTERSYSTEM'
#
#			#CI type
#			w_sheet2_cis['CL' +str(4+i)]='BMC_COMPUTERSYSTEM'
#			##Product type
#			w_sheet2_cis['CD' +str(4+i)]='Hardware'
#			####Supported
#			w_sheet2_cis['AC' +str(4+i)]='Yes'
#			###CI Name
#			w_sheet2_cis['D' +str(4+i)]=cis.filter(regex=re.compile('CI NAME',re.IGNORECASE)).iloc[:,0].values[i]
#
#			##CI ID
#			#w_sheet2_cis.write(4+i,11,cis.filter(regex=re.compile('CI ID',re.IGNORECASE)).iloc[:,0][i])
#			###CI Description
#			w_sheet2_cis['G' +str(4+i)]=cis.filter(regex=re.compile('CI Description',re.IGNORECASE)).iloc[:,0].values[i]
#
#			####Status
#			if (cis.filter(regex=re.compile('STATUS',re.IGNORECASE)).iloc[:,0]=='').any():
#				w_sheet2_cis['AP' +str(4+i)]='Deployed'
#			else:
#				w_sheet2_cis['AP' +str(4+i)]=cis.filter(regex=re.compile('STATUS',re.IGNORECASE)).iloc[:,0].values[i]		
#			####Tier 1
#			w_sheet2_cis['B' +str(4+i)]=cis['Tier 1'][i]
#			####Tier 2
#			w_sheet2_cis['H' +str(4+i)]=cis['Tier 2'][i]
#			####Tier 3
#			w_sheet2_cis['J' +str(4+i)]=cis['Tier 3'][i]
#			###Product Name
#			#w_sheet2_cis.write(4+i,13,cis['Product Name+'][i])
#			w_sheet2_cis['N' +str(4+i)]=cis.filter(regex=re.compile('Product Name',re.IGNORECASE)).iloc[:,0].values[i]
#
#			####Manufacturer
#			w_sheet2_cis['V' +str(4+i)]=cis['Manufacturer'][i]
#
#			####System Role
#			w_sheet2_cis['AN' +str(4+i)]=cis['System Role'][i]
#
#			###Priority
#			if (cis['Priority']=='').any():
#				w_sheet2_cis['BB' +str(4+i)]='PRIORITY_5'
#			else:
#				w_sheet2_cis['BB' +str(4+i)]=cis['Priority'][i].upper()
#
#			###Additional Information
#			w_sheet2_cis['BE' +str(4+i)]=cis['Additional Information'][i]
#			##
#			###Region
#			w_sheet2_cis['AL' +str(4+i)]=cis['Region'][i]
#			##
#			###Site Group
#			w_sheet2_cis['AR' +str(4+i)]=cis['Site Group'][i]
#			##
#			###Site
#			w_sheet2_cis['AT' +str(4+i)]=cis[cis.columns[~cis.columns.str.contains('Group',case=False)].tolist()].filter(regex=re.compile('SITE',re.IGNORECASE)).iloc[:,0].values[i]
#			##
#			###Tag Number
#			w_sheet2_cis['S' +str(4+i)]=cis['Tag Number'][i]
#
#			##
#			###Model Version
#			###w_sheet2_cis.write(4+i,19,cis[0].filter(regex=re.compile('Version',re.IGNORECASE)).iloc[:,0][i])
#			##
#			###DNS
#			w_sheet2_cis['AJ' +str(4+i)]=cis['DNS Host Name'][i]
#
#			##
#			###Domain
#			w_sheet2_cis['AQ' +str(4+i)]=cis['Domain'][i]
#
#
#			##
#			###num cells
#			if cis.columns.str.contains('cell',case=False).any():
#				w_sheet2_cis['DW' +str(4+i)]=cis.filter(regex=re.compile('cell',re.IGNORECASE)).iloc[:,0][i]
#			else:
#				print('No cells')
#		noam_cis.save(filename = NOAM_report + company +'_CIs_Noam_'+ dt.datetime.now().strftime("%Y-%m-%d %H-%M-%S") + '.xlsm')
#		#wb_cis.save(NOAM_report + company +'_cis_Noam.xls')
#	else:
#		None
#
